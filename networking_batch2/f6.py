from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = "ID"
ws["B1"] = "Name"

wb.save("employee.xls")